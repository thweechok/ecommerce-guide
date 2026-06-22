"""
แบ่งรายงานรายได้ TikTok Shop รายเดือน ออกเป็นไฟล์ Excel รายวัน
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• อ่านชีตแรกแบบ streaming (read_only=True) — ประหยัด RAM
• แยกออเดอร์ตามคอลัมน์ D «เวลาที่ชำระคำสั่งซื้อ» (YYYY/MM/DD)
• เขียนไฟล์ Excel แยกรายวันด้วย write_only workbook
• ห่อเป็น .zip ให้ดาวน์โหลดทีเดียว

รันด้วย:  streamlit run app.py
"""

import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
import zipfile
import io
import os
import re
from datetime import datetime

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ตั้งค่าหน้าเว็บ
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.set_page_config(
    page_title="แบ่งรายงาน TikTok Shop รายวัน",
    page_icon="🛒",
    layout="centered",
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Custom CSS — ธีมสไตล์ TikTok (สีเข้ม + accent สีชมพู-แดง)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<style>
/* ── ฟอนต์ภาษาไทย ── */
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+Thai:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] {
    font-family: 'Noto Sans Thai', 'Segoe UI', sans-serif;
}

/* ━━━━━━━━━━ Keyframe Animations ━━━━━━━━━━ */
@keyframes fadeSlideUp {
    from { opacity: 0; transform: translateY(28px); }
    to   { opacity: 1; transform: translateY(0); }
}
@keyframes fadeIn {
    from { opacity: 0; }
    to   { opacity: 1; }
}
@keyframes popIn {
    0%   { opacity: 0; transform: scale(.85); }
    60%  { opacity: 1; transform: scale(1.04); }
    100% { opacity: 1; transform: scale(1); }
}
@keyframes shimmer {
    0%   { background-position: -200% center; }
    100% { background-position: 200% center; }
}
@keyframes gradientShift {
    0%   { background-position: 0% 50%; }
    50%  { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}
@keyframes pulseGlow {
    0%, 100% { box-shadow: 0 4px 16px rgba(255,0,80,.3); }
    50%      { box-shadow: 0 6px 28px rgba(255,0,80,.55); }
}
@keyframes slideInLeft {
    from { opacity: 0; transform: translateX(-18px); }
    to   { opacity: 1; transform: translateX(0); }
}
@keyframes countPop {
    0%   { opacity: 0; transform: scale(.5); }
    70%  { transform: scale(1.1); }
    100% { opacity: 1; transform: scale(1); }
}
@keyframes borderGlow {
    0%, 100% { border-color: rgba(0,200,83,.25); box-shadow: 0 0 8px rgba(0,200,83,.1); }
    50%      { border-color: rgba(0,200,83,.5);  box-shadow: 0 0 18px rgba(0,200,83,.2); }
}
@keyframes borderGlowFail {
    0%, 100% { border-color: rgba(255,0,50,.3); box-shadow: 0 0 8px rgba(255,0,50,.1); }
    50%      { border-color: rgba(255,0,50,.6); box-shadow: 0 0 18px rgba(255,0,50,.25); }
}

/* ── Hero banner ── */
.hero {
    background: linear-gradient(135deg, #FF0050, #8B00FF, #00F2EA, #FF0050);
    background-size: 300% 300%;
    animation: gradientShift 6s ease infinite, fadeSlideUp .7s ease-out both;
    border-radius: 16px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
    text-align: center;
    box-shadow: 0 8px 32px rgba(255, 0, 80, .25);
    transition: box-shadow .3s ease, transform .3s ease;
}
.hero:hover {
    box-shadow: 0 12px 40px rgba(255, 0, 80, .35);
    transform: translateY(-2px);
}
.hero h1 {
    color: #fff;
    font-size: 1.75rem;
    font-weight: 700;
    margin: 0 0 .3rem;
    text-shadow: 0 2px 8px rgba(0,0,0,.3);
    animation: fadeSlideUp .7s ease-out .15s both;
}
.hero p {
    color: rgba(255,255,255,.88);
    font-size: .95rem;
    margin: 0;
    animation: fadeIn .8s ease-out .35s both;
}

/* ── Card wrapper ── */
.card {
    background: rgba(255,255,255,.04);
    border: 1px solid rgba(255,255,255,.08);
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    backdrop-filter: blur(8px);
    animation: fadeSlideUp .5s ease-out both;
    transition: border-color .3s ease, box-shadow .3s ease;
}
.card:hover {
    border-color: rgba(255,255,255,.15);
    box-shadow: 0 4px 20px rgba(0,0,0,.15);
}

/* ── สรุปรายวัน — ตาราง ── */
.summary-table {
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    border-radius: 10px;
    overflow: hidden;
    font-size: .92rem;
    animation: fadeSlideUp .5s ease-out both;
}
.summary-table thead th {
    background: rgba(255,0,80,.18);
    color: #FF6B9D;
    padding: .65rem 1rem;
    text-align: left;
    font-weight: 600;
    border-bottom: 1px solid rgba(255,0,80,.25);
}
.summary-table tbody tr {
    animation: slideInLeft .4s ease-out both;
    transition: background .25s ease, transform .2s ease;
}
.summary-table tbody tr:nth-child(1)  { animation-delay: .05s; }
.summary-table tbody tr:nth-child(2)  { animation-delay: .1s; }
.summary-table tbody tr:nth-child(3)  { animation-delay: .15s; }
.summary-table tbody tr:nth-child(4)  { animation-delay: .2s; }
.summary-table tbody tr:nth-child(5)  { animation-delay: .25s; }
.summary-table tbody tr:nth-child(6)  { animation-delay: .3s; }
.summary-table tbody tr:nth-child(7)  { animation-delay: .35s; }
.summary-table tbody tr:nth-child(8)  { animation-delay: .4s; }
.summary-table tbody tr:nth-child(9)  { animation-delay: .45s; }
.summary-table tbody tr:nth-child(10) { animation-delay: .5s; }
.summary-table tbody tr:nth-child(n+11){ animation-delay: .55s; }
.summary-table tbody td {
    padding: .55rem 1rem;
    border-bottom: 1px solid rgba(255,255,255,.05);
}
.summary-table tbody tr:hover {
    background: rgba(255,0,80,.08);
    transform: translateX(4px);
}
.summary-table tbody tr:last-child td {
    border-bottom: none;
}
.summary-table tfoot td {
    background: rgba(0,242,234,.08);
    color: #00F2EA;
    padding: .65rem 1rem;
    font-weight: 700;
    border-top: 2px solid rgba(0,242,234,.2);
    animation: fadeIn .5s ease-out .6s both;
}

/* ── Badge ตัวเลข ── */
.metric-row {
    display: flex;
    gap: 1rem;
    margin: 1rem 0;
}
.metric-box {
    flex: 1;
    background: rgba(255,255,255,.04);
    border: 1px solid rgba(255,255,255,.08);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    text-align: center;
    animation: popIn .5s ease-out both;
    transition: transform .25s ease, border-color .25s ease, box-shadow .25s ease;
}
.metric-box:first-child { animation-delay: .1s; }
.metric-box:last-child  { animation-delay: .25s; }
.metric-box:hover {
    transform: translateY(-3px);
    border-color: rgba(255,0,80,.3);
    box-shadow: 0 6px 20px rgba(255,0,80,.12);
}
.metric-box .label { font-size: .8rem; color: rgba(255,255,255,.5); margin-bottom: .3rem; }
.metric-box .value {
    font-size: 1.6rem;
    font-weight: 700;
    color: #FAFAFA;
    animation: countPop .6s ease-out .3s both;
}

/* ── ตรวจสอบ — match / mismatch ── */
.check-ok {
    background: rgba(0,200,83,.1);
    border: 1px solid rgba(0,200,83,.25);
    border-radius: 10px;
    padding: .8rem 1.2rem;
    color: #4ADE80;
    font-weight: 600;
    text-align: center;
    animation: popIn .5s ease-out both, borderGlow 2.5s ease-in-out infinite;
}
.check-fail {
    background: rgba(255,0,50,.1);
    border: 1px solid rgba(255,0,50,.3);
    border-radius: 10px;
    padding: .8rem 1.2rem;
    color: #FF4D6D;
    font-weight: 600;
    text-align: center;
    animation: popIn .5s ease-out both, borderGlowFail 1.5s ease-in-out infinite;
}

/* ── ปุ่มดาวน์โหลด — pulse glow ── */
div.stDownloadButton > button {
    background: linear-gradient(135deg, #FF0050 0%, #C70039 100%) !important;
    color: #fff !important;
    font-weight: 600 !important;
    font-size: 1.05rem !important;
    padding: .75rem 2rem !important;
    border-radius: 10px !important;
    border: none !important;
    box-shadow: 0 4px 16px rgba(255,0,80,.3) !important;
    transition: all .25s ease !important;
    animation: pulseGlow 2.5s ease-in-out infinite, fadeSlideUp .5s ease-out both !important;
}
div.stDownloadButton > button:hover {
    transform: translateY(-2px) scale(1.02) !important;
    box-shadow: 0 8px 30px rgba(255,0,80,.5) !important;
    animation: none !important;
}
div.stDownloadButton > button:active {
    transform: translateY(0) scale(.98) !important;
}

/* ── ปุ่ม primary ทั่วไป ── */
div.stButton > button[kind="primary"] {
    transition: all .25s ease !important;
}
div.stButton > button[kind="primary"]:hover {
    transform: translateY(-1px) scale(1.01) !important;
    box-shadow: 0 6px 24px rgba(255,0,80,.35) !important;
}
div.stButton > button[kind="primary"]:active {
    transform: translateY(0) scale(.98) !important;
}

/* ── Selectbox & input hover ── */
div[data-baseweb="select"] {
    transition: box-shadow .25s ease;
}
div[data-baseweb="select"]:hover {
    box-shadow: 0 0 0 1px rgba(255,0,80,.3);
}

/* ── File uploader ── */
section[data-testid="stFileUploader"] {
    animation: fadeSlideUp .5s ease-out .1s both;
}

/* ── ตัวอย่างชื่อไฟล์ shimmer ── */
.file-preview b {
    background: linear-gradient(90deg, #FF6B9D 0%, #FFD700 50%, #FF6B9D 100%);
    background-size: 200% auto;
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    animation: shimmer 3s linear infinite;
}

/* ── ซ่อน Streamlit default ── */
#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ฟังก์ชันช่วย
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# regex จับวันที่ YYYY/MM/DD (อาจมีเวลาต่อท้าย)
_DATE_RE = re.compile(r"^(\d{4})/(\d{2})/(\d{2})")

# ค่าคงที่สำหรับแถวที่ไม่มีวันที่
_NO_DATE = "__no_date__"


def _extract_date(value):
    """แยกวันที่จากค่าเซลล์ → คืน 'YYYY/MM/DD' หรือ None"""
    if value is None:
        return None
    # กรณีเป็น datetime object (บาง Excel เก็บเป็นวันที่จริง)
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    # กรณีเป็นข้อความ
    text = str(value).strip()
    m = _DATE_RE.match(text)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return None


def _make_row_cells(ws, values, ncols):
    """สร้าง list ของ WriteOnlyCell พร้อม format
    — คอลัมน์ A (index 0) = หมายเลขคำสั่งซื้อ → บังคับเป็นข้อความเสมอ
      เพื่อป้องกัน Excel แปลงเลข 18 หลักเป็น scientific notation"""
    cells = []
    for i in range(ncols):
        v = values[i] if i < len(values) else None
        cell = WriteOnlyCell(ws, value=v)
        # คอลัมน์ A: บังคับ text format
        if i == 0 and v is not None:
            cell.value = str(v)
            cell.number_format = "@"
        cells.append(cell)
    return cells


def _build_summary_html(summary, total_orders):
    """สร้างตาราง HTML สำหรับแสดงสรุปรายวัน"""
    rows_html = ""
    for item in summary:
        rows_html += (
            f'<tr><td>{item["date"]}</td>'
            f'<td style="text-align:right;">{item["count"]:,}</td></tr>\n'
        )
    return f"""
    <table class="summary-table">
        <thead><tr><th>📅 วันที่</th><th style="text-align:right;">จำนวนออเดอร์</th></tr></thead>
        <tbody>{rows_html}</tbody>
        <tfoot><tr><td>รวมทั้งหมด</td><td style="text-align:right;">{total_orders:,}</td></tr></tfoot>
    </table>
    """


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Hero Banner
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<div class="hero">
    <h1>🛒 แบ่งรายงานรายได้ E-Commerce เป็นรายวัน</h1>
    <p>อัปโหลดไฟล์ .xlsx → เลือกแพลตฟอร์มและประเภทรายงาน → ระบบแยกออเดอร์ตามวันที่ชำระ → ดาวน์โหลด .zip ที่มี Excel แยกทีละวัน</p>
</div>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  อัปโหลดไฟล์
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
uploaded = st.file_uploader(
    "📁 เลือกไฟล์ .xlsx (export จาก E-Commerce platform)",
    type=["xlsx"],
    help="รองรับไฟล์สูงสุด 300 MB  ·  ระบบจะอ่านเฉพาะชีตแรก",
)

if uploaded is None:
    st.markdown(
        '<p style="text-align:center; color:rgba(255,255,255,.4); margin-top:2rem;">'
        "👆 เลือกไฟล์เพื่อเริ่มต้นใช้งาน</p>",
        unsafe_allow_html=True,
    )
    st.stop()

# ── เมื่อไฟล์เปลี่ยน → ล้างผลลัพธ์เก่า ──
_file_id = f"{uploaded.name}|{uploaded.size}"
if st.session_state.get("_file_id") != _file_id:
    st.session_state.pop("result", None)
    st.session_state["_file_id"] = _file_id

st.markdown(
    f'<p style="color:rgba(255,255,255,.5); font-size:.85rem;">'
    f"📄 {uploaded.name} — {uploaded.size / 1_048_576:.1f} MB</p>",
    unsafe_allow_html=True,
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ตั้งชื่อไฟล์ — เลือกแพลตฟอร์ม + ประเภทรายงาน
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("### 🏷️ ตั้งชื่อไฟล์ผลลัพธ์")

_CUSTOM = "อื่นๆ (พิมพ์เอง)"

# ── Dropdown 1: แพลตฟอร์ม ──
_PLATFORMS = ["TikTok", "Shopee", "Lazada", _CUSTOM]
col_p1, col_p2 = st.columns([1, 1])
with col_p1:
    platform_choice = st.selectbox("แพลตฟอร์ม", _PLATFORMS, index=0)
with col_p2:
    if platform_choice == _CUSTOM:
        platform_custom = st.text_input("พิมพ์ชื่อแพลตฟอร์ม", placeholder="เช่น NocNoc")
    else:
        platform_custom = ""

# ── Dropdown 2: ประเภทรายงาน ──
_REPORT_TYPES = ["คำสั่งซื้อ", "รายรับของฉัน", "รับเงินเข้าธนาคาร", "ค่าใช้จ่ายแพลตฟอร์ม", _CUSTOM]
col_r1, col_r2 = st.columns([1, 1])
with col_r1:
    report_choice = st.selectbox("ประเภทรายงาน", _REPORT_TYPES, index=0)
with col_r2:
    if report_choice == _CUSTOM:
        report_custom = st.text_input("พิมพ์ประเภทรายงาน", placeholder="เช่น ค่าโฆษณา")
    else:
        report_custom = ""

# ── สร้าง prefix สำหรับชื่อไฟล์ ──
_platform = platform_custom.strip() if platform_choice == _CUSTOM else platform_choice
_report = report_custom.strip() if report_choice == _CUSTOM else report_choice

if not _platform:
    st.warning("⚠️ กรุณาระบุชื่อแพลตฟอร์ม")
    st.stop()
if not _report:
    st.warning("⚠️ กรุณาระบุประเภทรายงาน")
    st.stop()

# prefix ที่จะใช้นำหน้าชื่อไฟล์ เช่น "TikTok คำสั่งซื้อ"
file_prefix = f"{_platform} {_report}"

# ── แสดงตัวอย่างชื่อไฟล์ ──
st.markdown(
    f'<p style="color:rgba(255,255,255,.45); font-size:.82rem;">'
    f"📝 ตัวอย่างชื่อไฟล์: <b style='color:#FF6B9D'>{file_prefix} 2026-05-31.xlsx</b></p>",
    unsafe_allow_html=True,
)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ปุ่มแบ่งวัน + ประมวลผล
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if st.button("🔄 เริ่มแบ่งวัน", type="primary", use_container_width=True):
    try:
        with st.status("⏳ กำลังประมวลผล…", expanded=True) as status:

            # ── ขั้นตอน 1: เปิดไฟล์ด้วย read_only ──
            st.write("📂 เปิดไฟล์ต้นฉบับ…")
            uploaded.seek(0)  # รีเซ็ต file pointer กรณีกดซ้ำ
            wb_in = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)

            if not wb_in.sheetnames:
                wb_in.close()
                st.error("❌ ไฟล์นี้ไม่มีชีตข้อมูลเลย")
                st.stop()

            ws_in = wb_in.worksheets[0]
            st.write(f'📋 ใช้ชีต: «{ws_in.title}»  ({len(wb_in.sheetnames)} ชีตทั้งหมด — ข้ามชีตอื่น)')

            # ── ขั้นตอน 2: อ่าน header (แถวที่ 1) ──
            row_iter = ws_in.iter_rows()
            try:
                header = [c.value for c in next(row_iter)]
            except StopIteration:
                wb_in.close()
                st.error("❌ ชีตแรกไม่มีข้อมูลเลย (ไม่พบแถว header)")
                st.stop()

            ncols = len(header)
            st.write(f"📊 พบ {ncols} คอลัมน์")

            # ── ขั้นตอน 3: streaming อ่านทีละแถว → เขียนลง write_only workbook ทันที ──
            #    ใช้ write_only=True เพื่อให้ openpyxl flush ข้อมูลลง temp file อัตโนมัติ
            #    แทนที่จะเก็บทั้งหมดใน RAM
            st.write("⏳ กำลังอ่านและแยกข้อมูลตามวันที่ชำระ…")

            books: dict[str, list] = {}  # date_key → [Workbook, Worksheet, row_count]
            total = 0
            progress_placeholder = st.empty()

            for row in row_iter:
                vals = tuple(c.value for c in row)

                # ข้ามแถวที่ว่างสนิท (ทุกเซลล์เป็น None)
                if all(v is None for v in vals):
                    continue

                # ดึงวันที่จากคอลัมน์ D (index 3)
                raw_date = vals[3] if len(vals) > 3 else None
                date_key = _extract_date(raw_date)
                bucket = date_key if date_key else _NO_DATE

                # สร้าง write_only workbook ใหม่สำหรับวันที่ยังไม่เคยเจอ
                if bucket not in books:
                    owb = Workbook(write_only=True)
                    ows = owb.create_sheet("รายละเอียดคำสั่งซื้อ")
                    ows.append(header)  # เขียน header เป็นแถวแรก
                    books[bucket] = [owb, ows, 0]

                # เขียนแถวข้อมูลลง workbook ของวันนั้น
                entry = books[bucket]
                entry[1].append(_make_row_cells(entry[1], vals, ncols))
                entry[2] += 1
                total += 1

                # แสดงความคืบหน้าทุก 5,000 แถว
                if total % 5_000 == 0:
                    progress_placeholder.text(f"  อ่านแล้ว {total:,} แถว …")

            wb_in.close()
            progress_placeholder.text(f"  อ่านครบ {total:,} แถว ✓")

            # ตรวจว่ามีข้อมูลหรือไม่
            if total == 0:
                st.warning("⚠️ ไม่พบแถวข้อมูล (มีเฉพาะ header)")
                # ปิด workbook ที่เปิดค้างทั้งหมด (ถ้ามี)
                for owb, _, _ in books.values():
                    owb.close()
                st.stop()

            # ── ขั้นตอน 4: สร้างไฟล์ ZIP ──
            st.write("📦 กำลังบีบอัดเป็น ZIP…")

            # เรียงลำดับวันที่ → «ไม่ระบุ» ไว้ท้ายสุด
            sorted_keys = sorted(k for k in books if k != _NO_DATE)
            if _NO_DATE in books:
                sorted_keys.append(_NO_DATE)

            zip_buf = io.BytesIO()
            summary: list[dict] = []
            written = 0

            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for key in sorted_keys:
                    owb, _, count = books[key]

                    # ตั้งชื่อไฟล์: "[prefix] 2026-05-31.xlsx"
                    if key == _NO_DATE:
                        fname = f"{file_prefix} ไม่ระบุวันที่ชำระ.xlsx"
                        display = "ไม่ระบุวันที่"
                    else:
                        fname = f"{file_prefix} {key.replace('/', '-')}.xlsx"
                        display = key

                    # บันทึก workbook → BytesIO → ZIP
                    fbuf = io.BytesIO()
                    owb.save(fbuf)
                    owb.close()
                    zf.writestr(fname, fbuf.getvalue())
                    fbuf.close()

                    written += count
                    summary.append({"date": display, "count": count})

            status.update(label="✅ ประมวลผลเสร็จสิ้น!", state="complete")

        # ── เก็บผลลัพธ์ใน session_state เพื่อให้คงอยู่หลัง rerun ──
        st.session_state["result"] = {
            "zip_data": zip_buf.getvalue(),
            "zip_name": f"{file_prefix}_รายวัน.zip",
            "summary": summary,
            "total": total,
            "written": written,
            "n_files": len(summary),
        }

    # ── จัดการข้อผิดพลาดต่างๆ ──
    except openpyxl.utils.exceptions.InvalidFileException:
        st.error("❌ ไม่สามารถเปิดไฟล์ได้ — กรุณาตรวจสอบว่าเป็นไฟล์ .xlsx ที่ถูกต้อง (ไม่ใช่ .xls หรือ .csv)")
    except zipfile.BadZipFile:
        st.error("❌ ไฟล์เสียหายหรือไม่ใช่รูปแบบ .xlsx — ลองดาวน์โหลดไฟล์ใหม่จาก TikTok Shop")
    except PermissionError:
        st.error("❌ ไม่มีสิทธิ์เข้าถึงไฟล์ — กรุณาปิดไฟล์ที่เปิดอยู่ใน Excel แล้วลองใหม่")
    except MemoryError:
        st.error("❌ หน่วยความจำไม่เพียงพอ — ไฟล์ใหญ่เกินไปสำหรับเครื่องนี้")
    except Exception as exc:
        st.error(f"❌ เกิดข้อผิดพลาดที่ไม่คาดคิด: {exc}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  แสดงผลลัพธ์ (จาก session_state — คงอยู่แม้กดดาวน์โหลดแล้ว rerun)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if "result" in st.session_state:
    r = st.session_state["result"]
    st.divider()

    # ── (1) ตารางสรุปรายวัน ──
    st.markdown(f"### 📋 สรุปจำนวนออเดอร์รายวัน — {r['n_files']} ไฟล์")
    st.markdown(
        _build_summary_html(r["summary"], r["written"]),
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── (2) ตรวจสอบความถูกต้อง ──
    st.markdown("### 🔍 ตรวจสอบความถูกต้อง")
    st.markdown(
        f"""
        <div class="metric-row">
            <div class="metric-box">
                <div class="label">แถวต้นฉบับ (ไม่รวม header)</div>
                <div class="value">{r['total']:,}</div>
            </div>
            <div class="metric-box">
                <div class="label">แถวที่เขียนออกรวม</div>
                <div class="value">{r['written']:,}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if r["total"] == r["written"]:
        st.markdown(
            f'<div class="check-ok">✅ จำนวนแถวตรงกัน ({r["total"]:,} แถว) — ข้อมูลครบถ้วน</div>',
            unsafe_allow_html=True,
        )
    else:
        diff = abs(r["total"] - r["written"])
        st.markdown(
            f'<div class="check-fail">❌ ไม่ตรงกัน! ต้นฉบับ {r["total"]:,} แถว '
            f'แต่เขียนออก {r["written"]:,} แถว — ขาดหาย {diff:,} แถว</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── (3) ปุ่มดาวน์โหลด ZIP ──
    st.download_button(
        label=f"⬇️ ดาวน์โหลด {r['zip_name']}",
        data=r["zip_data"],
        file_name=r["zip_name"],
        mime="application/zip",
        type="primary",
        use_container_width=True,
    )

    st.markdown(
        '<p style="text-align:center; color:rgba(255,255,255,.3); font-size:.78rem; margin-top:.5rem;">'
        "ไฟล์ ZIP ภายในมี Excel แยกทีละวัน พร้อมใช้งาน</p>",
        unsafe_allow_html=True,
    )
